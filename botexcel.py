from openpyxl import *
import random
import datetime
from telegram import *
from telegram.ext import *
token = "2096295994:AAEagJoKQ6UvxzLVBzuSg44Q9-aAXXeGRTw"
bot = Bot(token)
updater = Updater(token, use_context=True)
dispatcher: Dispatcher = updater.dispatcher

d = datetime.datetime.now()
wb = Workbook()
i = 36010101
wb = load_workbook(str(i)+".xlsx")
ws = wb.active
name_bakhshha = ["مرکزی", "کهک", "سلفچگان", "دستجرد", "جعفریه"]
name_chaha_jafirieh = [
    'چاه دولت آباد',
    'چاه سعدآباد',
    'چاه شهید علی محمدی 1',
    'چاه شهید علی محمدی 2',
    'چاه شهید علی محمدی 3',
    'چاه شهید علی محمدی 4',
    'چاه کلاغ نشین',
    'چاه ناحیه صنعتی طغرود',
    'چاه شهری - اداره 1',
    'چاه شهری - بندر عباس 3',
    'چاه شهری - پارکینگ 2',
    'چاه شهری - شوری 4',
    'چاه اسفید',
    'چاه آغلک',
    'چاه چمانک و الگان',
    'چاه قاهان',
    'چاه کهندان',
    'چاه مهرزمین',
    'چاه موشکیه',
    'چاه وسفونجرد'
    ]
name_chaha_dastjerd = [
    'چاه احمد آباد',
    'چاه آمره 1',
    'چاه آمره 2',
    'چاه جریک آغاج',
    'چاه جوزه',
    'چاه حسن آباد 2',
    'چاه رستگان',
    'چاه زیزگان',
    'چاه سرخده',
    'چاه سفت',
    'چاه سفیدآله',
    'چاه سناوند 2',
    'چاه سناوند 3',
    'چاه طینوج 1',
    'چاه طینوج 2',
    'چاه عیسی آباد',
    'چاه کندرود 2',
    'چاه کندرود 3',
    'چاه گیو 1',
    'چاه گیو 2',
    'چاه منصورآباد',
    'چاه موجان',
    'چاه نایه',
    'چاه ورزنه 1',
    'چاه ورزنه 2',
    'چاه وشاره',
    'چاه شهری - رودخانه 1',
    'چاه شهری - نوح آباد 2',
    'چاه شهری -3'
    ]
name_chaha_salafchegan = [
    'چاه باغ یک',
    'چاه دیزیجان',
    'چاه راهجرد',
    'چاه راهدارخانه راهجرد',
    'چاه زواریان 2',
    'چاه زواریان 3',
    'چاه ساریه خاتون',
    'چاه سنجگان',
    'چاه علی آباد نیزار',
    'چاه عنایت بیگ',
    'چاه قراسو',
    'چاه قلعه چم',
    'چاه مجتمع قاضی 1',
    'چاه مجتمع قاضی 2',
    'چاه مجتمع گلستانه',
    'چاه یکه باغ 1',
    'چاه یکه باغ 2',
    'چاه شهری - کنار پمپ بنزین 2',
    'چاه شهری - کنار رودخانه 1'
    ]
name_chaha_kahak = [
    'چاه ابرجس',
    'چاه امامزاده اسماعیل 1',
    'چاه امامزاده اسماعیل 2',
    'چاه تیره',
    'چاه خاوه',
    'چاه دستگرد',
    'چاه سیرو 1',
    'چاه سیرو 2',
    'چاه صرم 1',
    'چاه قبادبزن 1',
    'چاه قبادبزن 2',
    'چاه قبادبزن حاج اسماعیل 5',
    'چاه قبادبزن قاسمی 6',
    'چاه کرمجگان 1',
    'چاه کرمجگان 2',
    'چاه میم',
    'چاه ورجان',
    'چاه وشنوه 1',
    'چاه وشنوه 2',
    'چاه ونارچ 1',
    'چاه ویریچ',
    'چاه شهری - آبنو 1',
    'چاه شهری - باغ آهو 2',
    ]
name_chaha_markazi = [
    'چاه پادگان خیبر',
    'چاه شوراب',
    'چاه صیدآباد',
    'چاه عباس آباد',
    'چاه قمرود',
    'چاه کاج',
    'چاه ملک قلعه',
    'چاه والیجرد',
    'چاه دامشهر',
    'چاه حاجی آباد آغا'
    ]
electropump = [
    "مشکلی ندارد",
    "بازدید های دوره ای انجام شده و موردخاصی مشاهده نگردیده است",
    "درمواقع لزوم توسط نیروهای پیمانکار روشن و خاموش می شود.",
    "به صورت دستی روشن و خاموش می شود",
    "توسط پیمانکار واحد نت تجهیز ارزیابی و کنترل می شود",
    "پیمانکار واحد نت پارامترهای فشار،دبی و ... را ماهیانه کنترل و اعلام می کند"
    ]
tablo_bargh = [
 "مشکلی ندارد",
 "بازدیدهای دوره ای انجام شده و مورد خاصی مشاهده نشده است",
 "در چک لیست تحویلی از واحدنت اطلاعات دقیق عملکردی تابلوبرق ثبت واعلام می شود",
    ]
khat_raneh = [
    "مشکلی ندارد"
    ] 
khamoosh_roshan_electropump = [
   '''
    در مواقع عدم نیاز آبی و همچنین اختلال شبکه برق به طور دستی
     الکتروپمپ خاموش شده و در صورت رفع مشکل به طور دستی روشن می شود.
    '''
    ]
tahvilgiri_tasisat_chah = [
    "از زمان راه ا ندازی چاه انجام شده است"
 ]
bardasht_maghadir = [
 "توسط پیمانکار پارامترها برداشت شده و در چک لیست واحد نت اعلام می شود"
    ]
alodagi_seil = [
    "در زمان مذکور انجام می شود"
        ]
nasht_roghan = [
    "نشت روغن :ندارد"
    ]
nasht_roghan = [
    "نشت روغن :ندارد"
    ]
seda = [
    "صدا و لرزش غیر عادی پمپ، شیرآلات و اتصالات :ندارد"
    ]
fuse = [
    "فیوزهای کات اوت ترانس :ندارد"
    ]
bo = [
    "بوی سوختگی و غیرعادی :ندارد"
    ]
cheragh_signal = [
    "چراغ های سیگنال  :ندارد"
    ]
etesalat = [
    "اتصالات : مشکلی ندارد"
    ]
dama = [
    "دمای محیط و تهویه :مشکلی ندارد"
    ]
andazegiri = [
    "اندازه گیری و ثبت آمپر و ولتاژ : در دوره های ماهیانه انجام می شود"
    ]
zaher = [
    "ظاهر تابلو:به طور دوره ای نظافت شده و تمیز میشود."
    ]
nezafat_tablo = [
    "نظافت تابلو:به طور دوره ای نظافت می شود"
    ]
zaher_roshanaee = ["ندارد"]
ensheabat_bargh = ["سه فاز است"]
moshahede_gasht_zani = [
    " توسط نیروهای پیمانکار گشت زنی می شود"
    ]
gasht_hadese = [
    "به دلیل بازبودن محوطه احتمال سرقت و خرابکاری بالاست"
    ]
serghat_darb_shode = [
    "به سرقت رفته است"
    ]
serghat_darb_nashode = [
    "خیر"
    ]
kharabi_hozche, khatarat_seil, ghrire_motearaf = [
    "خیر",
    "خیر",
    "خیر",
    ]
dastoraolamal_boharan = [

    "بستگی به موقعیت (عادی و بحران) طبق پروتکل اعلامی انجام می شود."
    ]
keyboard_1 = [
            [InlineKeyboardButton(
             "فایل خام را دارم ",
             callback_data="fil")]]
reply_mark_1 = InlineKeyboardMarkup(keyboard_1)


def messagetoUs(update: Update, context: CallbackContext):
    print(update)
    try:
        bot.send_message(
            chat_id="202910393",
            text='''
            این ربات حاوی آموزش رایگان مطالبی جهت پپیشگیری و درمان کرونا براساس احادیث اسلامی تهیه شده است.مطالب ارائه شده به شرط نخوردن هرگونه دم کرده،بخور، معجون و ... تضمین می گردد . جهت کسب اطلاعات بیشتر می توانید از قسمت ارتباط با ما در ارتباط باشید.
                    ''',
            reply_markup=reply_mark_1,
            )
        ws['d8'] = update.message.text
    except Exception as e:
        pass


def query_btns(update: Update, context: CallbackContext):
    query: CallbackQuery = update.callback_query
    if query.data == "fil":
        bot.send_message(
            chat_id=update.effective_message.chat_id,
            text='''
        55555
            '''
            )
    for x in range(len(name_bakhshha)):
        for c in range(len(electropump)):
            ws['a8'] = random.choice(electropump)
        for c in range(len(tablo_bargh)):
            ws['b8'] = random.choice(tablo_bargh)
        for c in range(len(khat_raneh)):
            ws['c8'] = random.choice(khat_raneh)
        for c in range(len(khamoosh_roshan_electropump)):
            ws['b12'] = random.choice(khamoosh_roshan_electropump)
        for c in range(len(khamoosh_roshan_electropump)):
            ws['b14'] = random.choice(khamoosh_roshan_electropump)
        for c in range(len(tahvilgiri_tasisat_chah)):
            ws['b13'] = random.choice(tahvilgiri_tasisat_chah)
        for c in range(len(bardasht_maghadir)):
            ws['b15'] = random.choice(bardasht_maghadir)
        for c in range(len(bardasht_maghadir)):
            ws['b16'] = random.choice(bardasht_maghadir)
        for c in range(len(alodagi_seil)):
            ws['b17'] = random.choice(alodagi_seil)
        for c in range(len(nasht_roghan)):
            ws['a20'] = random.choice(nasht_roghan)
        for c in range(len(seda)):
            ws['b20'] = random.choice(seda)
        for c in range(len(fuse)):
            ws['c20'] = random.choice(fuse)
        for c in range(len(zaher)):
            ws['d21'] = random.choice(zaher)
        for c in range(len(bo)):
            ws['e20'] = random.choice(bo)
        for c in range(len(cheragh_signal)):
            ws['a21'] = random.choice(cheragh_signal)
        for c in range(len(etesalat)):
            ws['b21'] = random.choice(etesalat)
        for c in range(len(dama)):
            ws['c21'] = random.choice(dama)
        for c in range(len(nezafat_tablo)):
            ws['d20'] = random.choice(nezafat_tablo)
        for c in range(len(andazegiri)):
            ws['b24'] = random.choice(andazegiri)
        for c in range(len(zaher_roshanaee)):
            ws['b24'] = random.choice(zaher_roshanaee)
        for c in range(len(ensheabat_bargh)):
            ws['b25'] = random.choice(ensheabat_bargh)
        for c in range(len(dastoraolamal_boharan)):
            ws['a36'] = random.choice(dastoraolamal_boharan)
            ws['f12'], ws['f13'], ws['f14'], ws['f15'], ws['f16'], ws['f17'] = "بلی", "بلی", "بلی", "بلی", "بلی", "بلی"
        if name_bakhshha[x] == "جعفریه":
            for name in range(len(name_chaha_jafirieh)):
                ws['a5'] = "آدرس چاه :"+name_chaha_jafirieh[name]
                ws['d3'] = "اسامی افراد ثبت کننده :"+"آقای محمودآبادی"
                b = datetime.timedelta(
                    days=226895+random.randint(0, 10) +
                    random.randint(400, 1000)/1000)
                ws['b4'] = (d-b).strftime("%Y-%m-%d , %H,%M'")
                wb.save(
                    filename=str(name_chaha_jafirieh[name])
                    + " "+str(name_bakhshha[x])+str(i) +
                    ".xlsx")

        elif name_bakhshha[x] == "دستجرد":
            for name in range(len(name_chaha_dastjerd)):
                ws['a5'] = "آدرس چاه :"+name_chaha_dastjerd[name]
                ws['d3'] = "اسامی افراد ثبت کننده :"+"آقای دینی"
                b = datetime.timedelta(
                    days=226895+random.randint(0, 10) +
                    random.randint(400, 1000)/1000)
                ws['b4'] = (d-b).strftime("%Y-%m-%d , %H,%M'")
                wb.save(
                    filename=str(name_chaha_dastjerd[name])
                    + " "+str(name_bakhshha[x])+str(i)+".xlsx")

        elif name_bakhshha[x] == "سلفچگان":
            for name in range(len(name_chaha_salafchegan)):
                ws['a5'] = "آدرس چاه :"+name_chaha_salafchegan[name]
                ws['d3'] = "اسامی افراد ثبت کننده :"+"آقای شیری"
                b = datetime.timedelta(
                    days=226895+random.randint(0, 10) +
                    random.randint(400, 1000)/1000)
                ws['b4'] = (d-b).strftime("%Y-%m-%d , %H,%M'")
                wb.save(
                    filename=str(name_chaha_salafchegan[name])
                    + " "+str(name_bakhshha[x])+str(i)+".xlsx")

        elif name_bakhshha[x] == "کهک":
            for name in range(len(name_chaha_kahak)):
                ws['a5'] = "آدرس چاه :"+name_chaha_kahak[name]
                ws['d3'] = "اسامی افراد ثبت کننده :"+"آقای کامی"
                b = datetime.timedelta(
                    days=226895+random.randint(0, 10) +
                    random.randint(400, 1000)/1000)
                ws['b4'] = (d-b).strftime("%Y-%m-%d , %H,%M'")
                wb.save(
                    filename=str(name_chaha_kahak[name])
                    + " "+str(name_bakhshha[x])+str(i)+".xlsx")

        elif name_bakhshha[x] == "مرکزی":
            for name in range(len(name_chaha_markazi)):
                ws['a5'] = "آدرس چاه :"+name_chaha_markazi[name]
                ws['d3'] = "اسامی افراد ثبت کننده :"+"آقای درویشی"
                b = datetime.timedelta(
                    days=226895+random.randint(0, 10) +
                    random.randint(400, 1000)/1000)
                ws['b4'] = (d-b).strftime("%Y-%m-%d , %H,%M'")
                wb.save(
                    filename=str(name_chaha_markazi[name])
                    + " "+str(name_bakhshha[x])+str(i)+".xlsx")


print(i)
dispatcher.add_handler(MessageHandler(Filters.text, messagetoUs))
dispatcher.add_handler(CallbackQueryHandler(query_btns))
updater.start_polling()
